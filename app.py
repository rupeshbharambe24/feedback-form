import importlib
import os
import openpyxl
from flask import Flask, render_template, request, redirect, url_for

app = Flask(__name__)
app.secret_key = os.urandom(24)  # Set a secret key for session management

def save_to_excel(data, sheet_name):
    # Load the existing workbook or create a new one
    try:
        wb = openpyxl.load_workbook('feedback_data.xlsx')
    except FileNotFoundError:
        wb = openpyxl.Workbook()

    # Select the active sheet (create a new one if it doesn't exist)
    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
    else:
        sheet = wb.create_sheet(title=sheet_name)

    # Check if the sheet is empty or if the title row is missing
    if sheet.max_row == 0 or sheet.cell(row=1, column=2).value is None:
        sheet.insert_rows(1)
        for idx, key in enumerate(data.keys(), start=2):
            sheet.cell(row=1, column=idx, value=key)

    # Find the next empty row
    next_row = sheet.max_row + 1

    # Insert data values
    for idx, value in enumerate(data.values(), start=2):
        sheet.cell(row=next_row, column=idx, value=value)

    wb.save('feedback_data.xlsx')


@app.route('/')
def index():
    return render_template('index.html')

@app.route('/start-curriculum')
def start_curriculum():
    return redirect('/curriculum-feedback')

@app.route('/curriculum-feedback')
def curriculum():
    return render_template('curriculum-feedback.html')

@app.route('/library-feedback', methods=['GET', 'POST'])
def library_feedback():
    if request.method == 'POST':
        return redirect('/success')
    return render_template('library-feedback.html') 

@app.route('/ambience-feedback', methods=['GET', 'POST'])
def ambience_feedback(): 
    if request.method == 'POST':
        return redirect('/success') 
    return render_template('ambience-feedback.html') 

@app.route('/faculty-feedback', methods=['GET', 'POST'])
def faculty_feedback(): 
    if request.method == 'POST':
        return redirect('/success') 
    return render_template('faculty-feedback.html')

@app.route('/submit-curriculum-feedback', methods=['POST'])
def submit_curriculum_feedback():
    curriculum_data = request.form
    save_to_excel(curriculum_data, 'Curriculum Feedback')
    return redirect('/library-feedback')

@app.route('/submit-library-feedback', methods=['POST'])
def submit_library_feedback():
    library_data = request.form
    save_to_excel(library_data, 'Library Feedback')
    return redirect('/ambience-feedback') 

@app.route('/submit-ambience-feedback', methods=['POST'])
def submit_ambience_feedback():
    ambience_data = request.form
    save_to_excel(ambience_data, 'Ambience Feedback')
    return redirect('/faculty-feedback')
    
@app.route('/submit-faculty-feedback', methods=['POST'])
def submit_faculty_feedback():
    # Process form data to get branch and semester information
    branch = request.form.get('branch')
    semester = request.form.get('semester')

    # Call the desired function from faculty.py based on branch and semester
    if branch == 'AI&DS' and semester == '1':
        result = "Go and Study"
   # elif branch == 'AI&DS' and semester == '2':

    elif branch == 'AI&DS' and semester == '3':
       return redirect('/subjects/AIAI')
   # elif branch == 'AI&DS' and semester == '4':

   # elif branch == 'AI&DS' and semester == '5':

   # elif branch == 'AI&DS' and semester == '6':

   # elif branch == 'AI&DS' and semester == '7':

   # elif branch == 'AI&DS' and semester == '8':
        
    else:
        result = 'No matching function found'

    return result

#This is route for AI&DS 3rd Semester
@app.route('/subjects/AIAI')
def AIAI_page():
    return render_template('subjects/AIAI.html')

# Route to serve subject page CAOS
@app.route('/subjects/CAOS')
def CAOS_page():
    return render_template('subjects/CAOS.html')

# Route to serve subject page DLSP
@app.route('/subjects/DLSP')
def DLSP_page():
    return render_template('subjects/DLSP.html')

# Route to serve subject page DSAP
@app.route('/subjects/DSAP')
def DSAP_page():
    return render_template('subjects/DSAP.html')

# Route to serve subject page M3
@app.route('/subjects/M3')
def M3_page():
    return render_template('subjects/M3.html')

@app.route('/submit-aiai-feedback', methods=['POST'])
def submit_aiai_feedback():
    aiai_data = request.form
    save_to_excel(aiai_data, 'AIAI Feedback')
    return redirect('subjects/CAOS')

@app.route('/submit-caos-feedback', methods=['POST'])
def submit_caos_feedback():
    caos_data = request.form
    save_to_excel(caos_data, 'CAOS Feedback')
    return redirect('subjects/DSAP')

@app.route('/submit-dsap-feedback', methods=['POST'])
def submit_dsap_feedback():
    dsap_data = request.form
    save_to_excel(dsap_data, 'DSAP Feedback')
    return redirect('subjects/DLSP')

@app.route('/submit-dlsp-feedback', methods=['POST'])
def submit_dlsp_feedback():
    dlsp_data = request.form
    save_to_excel(dlsp_data, 'DLSP Feedback')
    return redirect('subjects/M3')

@app.route('/submit-m3-feedback', methods=['POST'])
def submit_m3_feedback():
    m3_data = request.form
    save_to_excel(m3_data, 'M3 Feedback')
    return redirect('/success')








@app.route('/success')
def success():
    return render_template('success.html')

if __name__ == '__main__':
    app.run(debug=True)
